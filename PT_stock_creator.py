from os import listdir, getcwd, system
from os.path import isfile, join
from pathlib import Path
from getpass import getpass
from re import findall
import requests
import warnings

from openpyxl import load_workbook
# from pandas import ExcelWriter, DataFrame  # TODO reduce import
import pandas as pd

from src.classes import Protein, attrib_dict

# TODO
# 1) decorators - get_file, task_start/end
# 2) don't add stocks already with IDs
# 3) update excel from LG
# 4) select sheets for update from/to LG
# 5) if sequence - calculate mass, len, pI, A0.1%


# API requests



def main():
    system('cls')
    
    global TOKEN
    global PB_ALL

    # LG Token acquisition
    test_mode = select_mode()
    TOKEN = get_token(test_mode=test_mode)
    
    PB_ALL = Path(r'P:\_research group folders\PT Proteins\_PT_stock_creator')

    warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

    # Main program loop
    while True:
        system('cls')
        menu_dict = {
            '1': ['Create new template file', 
                generate_template],
            '2': ['Add new protein records and stocks to Labguru',
                add_to_lg],
            '3': ['Update Labguru records and stocks from excel',
                update_to_lg],
            '4': ['Update excel records and stocks from Labguru',
                update_from_lg],
            '5': ['Create excel file for label printing',
                create_label_xlsx],
                }

        # Print menu
        print_menu(menu_dict)

        user_input = input('\nSelect task number: ')

        if user_input in menu_dict:
            system('cls')
            menu_dict[user_input][1]()
            system('pause')
        elif user_input.lower() == "q":
            break
        else:
            print('---< Wrong input. Try again >---'.center(80))
            system('pause')

# Helpers

def get_token(test_mode=False):
    """Acquires token from Labguru API

    Args:
        test_mode (bool, optional): Access test Labguru environment. Defaults to False.

    Returns:
       str: Token for API connection
    """
    print('\nEnter Labguru credentials: name (n.surname) and password')
    while True:

        if test_mode is True:
            name = 'LG_User7@purebiologics.com'
            password = 'LG_user7'
        else:
            name = input('Name: ').lower()
            password = str(getpass('Password: '))
        
        body = {'login': name + '@purebiologics.com',
                'password': password}
        
        message = '\nAcquiring authentication token'
        print(f'{message.ljust(73, ".")}', end='')
        
        try:
            session = requests.post('https://my.labguru.com/api/v1/sessions.json', body)
            token = session.json()['token']
            
            if token and token != '-1':
                print('SUCCESS')
                break
            else:
                print('..ERROR')
                print('Wrong login or password. Try again\n')
        except Exception as e:
            print(e)
            
    return token

def select_mode():
    """Selection of test mode for access to test Labguru environment

    Returns:
        bool: Test mode boolean flag
    """
    while True:
        mode_selection = input('Test mode (Y/N): ').lower()
        if 'y' in mode_selection:
            return True
        elif 'n' in mode_selection:
            return False
        else:
            print('---< Wrong input. Try again >---'.center(80))

def print_menu(menu_dict):
        print('/ PT Stock creator \\'.center(80, '_'), '\n')
        for k, v in menu_dict.items():
            print(f'{k:>2}) {v[0]}')
        print('\n Q - Quit')
        print(''.center(80, '_'))

# def task_start(file = None):
#     print("/ Task in progress \\".center(80, '_'), '\n', sep='')
#     if file is not None:
#         print(f'File: {file}\n')

# def task_end():
#     "Prints a bar after ending a task"
#     print("\n", "\\ Task completed /".center(80, '‾'), '\n', sep="")

def pball_connection(file):
    """Verifies if PB_all file is accessible

    Args:
        file (str): File name
    """
    while True:
        if isfile(join(PB_ALL, file)):
            break        
        else:
            print('---< Template file not found: check connection to PB_all >---'.center(80))
            system('pause')

def task_frame(func):
    def wrap(*args, **kwargs):

        print("/ Task in progress \\".center(80, '_'), '\n', sep='')
        result = func(*args, **kwargs)
        print("\n", "\\ Task completed /".center(80, '‾'), '\n', sep="")
        
        return result
    return wrap

def get_path_file(ext: str):
    """_summary_

    Args:
        ext (str): File format extension

    Raises:
        Exception: _description_

    Returns:
        mypath (str):   absolute path to directory of the file
        file (str):     name of selected file
    """
    mypath = getcwd()
    xlsx_list = scan_files(mypath, ext)
    file = choose_file(xlsx_list, ext)
    if file is None:
        raise Exception

    system('cls')
    print("/ Task in progress \\".center(80, '_'), '\n', sep='')
    print(f'\nFile: {file}\n')
    
    return mypath, file
        

def scan_files(path, ext):
    """Return list of all files in path with given extension

    Args:
        path (str): path to target directory
        ext (str):  file extension

    Returns:
        list: list of files with given extension
    """
    xlsx_list = [file for file in listdir(path) if check_filename(path, file, ext)]
    return xlsx_list

def check_filename(path, file: str, ext: str):
    """Return True if there is a non-temp file in given path with given extension else False

    Args:
        path (str): path to file location
        file (str): file name
        ext (str):  extension

    Returns:
        bool: True if non-temporary file exists in given location
    """
    is_file = isfile(join(path, file))
    is_ext = file.endswith(f'.{ext}')
    not_temp = not file.startswith('~$')
    conditions = [is_file, is_ext, not_temp]
    return all(conditions)

def choose_file(file_list: list, ext: str):
    print('Select file:\n',)
    
    xlsx_dict = {str(i): file for i, file in enumerate(file_list, 1)}    
    for i, file in xlsx_dict.items():
        print(f'{i.rjust(2)})', file.replace(f'.{ext}', ''), sep=" ")
        
    print('\n', 'Q - Return to menu')
    print(''.center(80, '_'), '\n', sep='')
    
    while True:
        file_i = input('File number: ').lower()
        if file_i in xlsx_dict:
            break
        elif file_i == 'q':
            return None
        else:
            print('---< Wrong input. Try again >---'.center(80))

    system('cls')
    
    return xlsx_dict[file_i]

def update_workbook(mypath, file, prot_list):
    while True:
        wb = load_workbook(join(mypath, file))
        try:
            for prot in prot_list:
                prot.update_sheet_prot(wb)
                prot.update_sheet_stocks(wb)
            wb.save(join(mypath, file))
        except PermissionError:
            print(f'---< Unable to save changes in {file} >---'.center(80))
            print('Close the file and continue'.center(80))
            system('pause')
        else:
            print(f'\nFile updated:\t\t {file}')
            break
        finally:
            wb.close()

def select_sheets(wb):
    print('\nSheets:')
    for i, sheet in enumerate(wb.sheetnames):
        print(f'   {i+1:>2d}) {sheet}')
    while True:
        try:
            selection = input('\nEnter sheet numbers or "a" to select all: ').lower()
            selection = list(findall(r"[\d]+|a", selection))
            if 'a' in selection:
                print()
                return wb.sheetnames
            elif (sheets := [wb.sheetnames[int(x)-1] for x in selection]):
                print()
                return sheets
            else:
                print('---< Wrong input. Try again >---'.center(80))
        except IndexError:
            print('---< Wrong input. Try again >---'.center(80))
            
def protein_from_sheet_gen(wb):
    """Yield Protein object from Workbook sheets

    Args:
        wb (openpyxl.Workbook): Workbook object from openpyxl

    Yields:
        Protein: Protein object based on data from Workbook sheet
    """
    ws_selection = select_sheets(wb)
    for ws_name in ws_selection:
        ws = wb[ws_name]
        protein_data = get_protein_data(ws)
        if protein_data.get('name'):
            yield Protein(protein_data, TOKEN.token)
        else:
            continue

def if_calc_params():
    """Return boolean flag if protein parameters should be calculated based on user input

    Returns:
        bool: True if parameters should be calculated
    """
    while True:
        resp = input('Calculate protein parameters based on the provided sequence? (Y/N): ').lower()
        if 'y' in resp:
            return True
        elif 'n' in resp:
            return False
        else:
            print('---< Wrong input. Try again >---'.center(80))


# TODO Move to Protein and refactor
def get_protein_data(ws) -> dict:

    protein_data = {'ws_name': ws.title}

    for attrib_name, value, *row in ws.iter_rows(min_row=0,
                            min_col=0,
                            # max_row=100,
                            max_col=2,
                            values_only=True):
        try:
            if attrib_name is not None:
                lg_name = attrib_dict.get(attrib_name, None)
                if lg_name is not None:
                    protein_data[lg_name] = value
        except KeyError as e:
            print(e)
    return protein_data


# 1) New template

@task_frame
def generate_template():
    """ Create copy of template file stored on PB_all """
    template_file = 'templates\\PT stock creator_template.xlsx'
    pball_connection(template_file)
    
    name = input('Name of the template file: ') + '.xlsx'
    mypath = getcwd()
    while name in listdir(mypath):
        print('---< A duplicate file name exists >---'.center(80))
        name = input('Select other name: ') + '.xlsx'
        
    system(f'copy "{join(PB_ALL, template_file)}" "./{name}" >nul')
    print(f'\nTemplate file {name} created')  

# 2) Add records and stocks

@task_frame
def add_to_lg():
    """
    
    """
    try:
        mypath, file = get_path_file('xlsx')
    except Exception:
        return

    prot_list = []
    prot_count = 0
    params_flag = if_calc_params() # TODO add to method
    
    wb = load_workbook(join(mypath, file), data_only=True)
    # ws_selection = select_sheets(wb)
    # for ws_name in ws_selection:
    #     ws = wb[ws_name]
    #     protein_data = get_protein_data(ws)
    #     protein = Protein(protein_data, TOKEN)
    for protein in protein_from_sheet_gen(wb):
        try:
            if params_flag:
                protein.calc_params()
            if protein.id is None:
                prot_count += protein.create_new_record()
        except Exception as e:  # TODO set exceptions and message
            print(e)
            pass
            
        protein.create_stocks(file)
        prot_list.append(protein)
        
        update_workbook(mypath, file, prot_list)

    stock_count = sum([len(prot.added_stocks) for prot in prot_list])

    wb.close()

    print(f'Protein entries created: {prot_count}')
    print(f'Stocks added:\t\t {stock_count:}\n')
    

# 3) Update existing records in Labguru
@task_frame
def update_to_lg():

    try:
        mypath, file = get_path_file('xlsx')
    except Exception:
        return
    
    wb = load_workbook(join(mypath, file), data_only=True)
    for protein in protein_from_sheet_gen(wb):
        if protein.id is not None:
            protein.update_lg_record()
        protein.update_lg_stocks(file) # TODO wb->ws ?
    wb.close()
    

# 4) Update excel file from Labguru records

@task_frame
def update_from_lg():

    try:
        mypath, file = get_path_file('xlsx')
    except Exception:
        return
    
    prot_list = []
    
    wb = load_workbook(join(mypath, file), data_only=True)
    for protein in protein_from_sheet_gen(wb):
        if protein.id is not None:
            protein.update_excel_record(wb)
            
        protein.update_excel_stocks(wb)
        prot_list.append(protein)
        
        update_workbook(mypath, file, prot_list)

# 5) Labels

@task_frame
def create_label_xlsx():
      
    mypath = getcwd()
    ext = 'xlsx'
    xlsx_list = scan_files(mypath, ext)
    file = choose_file(xlsx_list, ext)
    if file is None:
        return None
    
    wb = load_workbook(join(mypath, file), data_only=True)
    columns = {'Stock ID': int,
               'Stock name': str,
               'Concentration': float,
               'Stock volume': int,
               'Stock mass': int,
               'Box name': str,
               'Position': str}
    dfs = []
    label_file = file.replace('.xlsx', '_labels.xlsx')
    
    # TODO turn into generator
    
    for protein in protein_from_sheet_gen(wb):
        stock_df = protein.get_stock_df(file).astype(columns)
        stock_df['Concentration'] = stock_df['Concentration'].map(lambda x: f'{x:.3f}')
        dfs.append((protein.ws_name, stock_df.loc[:, columns.keys()]))
    
    while True:
        try:
            with pd.ExcelWriter(label_file) as writer:
                    for ws, df in dfs:
                        df.to_excel(writer,
                                    sheet_name=ws,
                                    index=True,
                                    header=False)
            break
        except PermissionError:
            print(f'---< Unable to save changes in {file} >---'.center(80))
            print('Close the file and continue'.center(80))
            system('pause')

    wb.close()
    print(f'File saved as {label_file}')


if __name__ == '__main__':
    main()
